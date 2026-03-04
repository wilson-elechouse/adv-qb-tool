#!/usr/bin/env python3
import argparse
import json
import re
import subprocess
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

HEADER_ALIASES = {
    "vendor": ["vendor", "vendor name", "supplier", "payee"],
    "bill_no": ["billing/soa no.", "billing soa no", "soa no", "billing no", "bill no", "feishu"],
    "request_no": ["request no.", "request no", "request number", "req no", "req#"],
    "billing_end_date": ["billing end date", "billing date", "end date"],
    "billing_start_date": ["billing start date", "start date"],
    "due_date": ["due date", "payment due date", "due"],
    "status": ["status", "approval status", "request status"],
    "location": ["location", "site", "branch"],
    "belongs_to": ["belongs to", "business unit", "bu", "class"],
    "project_type": ["project type", "project", "location mapping"],
    "reason": ["reason", "reasons for payment", "payment reason", "description reason"],
    "payment_details_01": ["payment detail 01", "payment details 01", "payment detail01", "payment details01"],
    "payment_details_02": ["payment detail 02", "payment details 02", "payment detail02", "payment details02"],
    "category": ["category", "account", "account name", "expense account"],
    "amount": ["amount", "line amount", "gross amount", "total amount"],
    "wht_rate": ["wht rate", "withholding tax rate", "withholding rate", "ewt rate"],
    "wht_amount": ["wht amount", "withholding tax amount", "ewt amount"],
    "vat_flag": ["vat in/ex", "vat in ex", "vat in", "vat ex", "vat type", "vat"],
}


def norm(s):
    return " ".join(str(s or "").strip().lower().replace("_", " ").split())


def to_date_text(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # handle excel-like datetime strings first
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    # fallback: split timestamp suffix if present
    if " " in s and len(s.split(" ")[0]) == 10:
        return s.split(" ")[0]
    return s


def to_period(v):
    if not v:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(v, fmt).strftime("%B %Y")
        except Exception:
            continue
    return ""


def detect_map(headers):
    idx = {}
    nheaders = [norm(h) for h in headers]
    for key, aliases in HEADER_ALIASES.items():
        for i, h in enumerate(nheaders):
            if any(a in h for a in aliases):
                idx[key] = i
                break
    return idx


def collect_section_text(rows, section_title):
    in_sec = False
    lines = []
    for r in rows:
        cells = [str(c or "").strip() for c in r]
        ncells = [norm(c) for c in cells]
        row_text = " | ".join([c for c in cells if c])

        if any(section_title in x for x in ncells):
            in_sec = True
            continue

        if in_sec and any(x.startswith("payment detail") and section_title not in x for x in ncells):
            break

        if in_sec and row_text:
            lines.append(row_text)

    return "\n".join(lines).strip()


def find_payment_detail_02_vendor(rows):

    in_pd2 = False
    for r in rows:
        cells = [str(c or "").strip() for c in r]
        ncells = [norm(c) for c in cells]

        if any("payment detail 02" in x for x in ncells):
            in_pd2 = True
            continue

        if in_pd2 and any(x.startswith("payment detail") and "02" not in x for x in ncells):
            in_pd2 = False

        if not in_pd2:
            continue

        for i, c in enumerate(cells):
            lc = norm(c)
            if "which supplier" not in lc:
                continue

            # Pattern A: same-cell key:value, e.g. "Which Supplier:Personal Supplier-Internal"
            m = re.search(r"which\s*supplier\s*[:\-]\s*([^|]+)", c, flags=re.IGNORECASE)
            if m and m.group(1).strip():
                return m.group(1).strip(), "payment_detail_02.which_supplier_inline"

            # Pattern B: same-cell key then value with spaces, e.g. "Which Supplier Personal Supplier-Internal"
            m2 = re.search(r"which\s*supplier\s+([^|]+)", c, flags=re.IGNORECASE)
            if m2 and m2.group(1).strip() and norm(m2.group(1).strip()) != "which supplier":
                return m2.group(1).strip(), "payment_detail_02.which_supplier_inline"

            # Pattern C: right-side cell value
            for j in range(i + 1, len(cells)):
                if str(cells[j]).strip():
                    return str(cells[j]).strip(), "payment_detail_02.which_supplier"

            return "", "payment_detail_02.which_supplier_found_but_empty"

    # Global fallback: scan any cell for inline "Which Supplier: <value>" pattern
    for r in rows:
        for c in [str(x or "").strip() for x in r]:
            m = re.search(r"which\s*supplier\s*[:\-]\s*([^|]+)", c, flags=re.IGNORECASE)
            if m and m.group(1).strip():
                return m.group(1).strip(), "global.which_supplier_inline_fallback"
            m2 = re.search(r"which\s*supplier\s+([^|]+)", c, flags=re.IGNORECASE)
            if m2 and m2.group(1).strip() and norm(m2.group(1).strip()) != "which supplier":
                return m2.group(1).strip(), "global.which_supplier_inline_fallback"

    return "", "payment_detail_02.which_supplier_not_found"


def parse_wht_rate_value(v):
    s = str(v or "").strip()
    if not s:
        return None
    m = re.search(r"(-?\d+(?:\.\d+)?)", s)
    if not m:
        return None
    n = float(m.group(1))
    if "%" in s or n > 1:
        n = n / 100.0
    if n < 0:
        return None
    return round(n, 6)


def extract_wht_rate(pd01_text, ai_cmd=None):
    text = str(pd01_text or "")
    if ai_cmd:
        payload = {
            "task": "extract_wht_rate",
            "input": {"payment_detail_01_text": text},
            "instruction": "Extract ONLY wht_rate as decimal number. Return strict JSON: {\"wht_rate\": number}. If missing, return 0."
        }
        try:
            p = subprocess.run(ai_cmd, input=json.dumps(payload, ensure_ascii=False), text=True, capture_output=True, shell=True)
            if p.returncode == 0 and (p.stdout or "").strip():
                obj = json.loads((p.stdout or "").strip())
                n = parse_wht_rate_value(obj.get("wht_rate", 0))
                if n is None:
                    return 0.0, "ai_parse_default_zero"
                return n, "ai_parse"
        except Exception:
            pass

    m = re.search(r"(?:2307\s*rate|wht\s*rate|withholding\s*tax\s*rate)\s*[:=]?\s*(-?\d+(?:\.\d+)?)\s*%?", text, flags=re.IGNORECASE)
    if m:
        n = parse_wht_rate_value(m.group(1) + ("%" if "%" in text[m.start():m.end()+2] else ""))
        return (n if n is not None else 0.0), "regex_parse"

    return 0.0, "not_found_default_zero"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--out", required=True, help="output json file path")
    ap.add_argument("--wht-ai-cmd", help="optional shell cmd for AI extraction of wht_rate")
    args = ap.parse_args()

    fp = Path(args.file)
    out_path = Path(args.out)

    wb = load_workbook(fp, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        out = {"ok": False, "error": "empty_sheet", "file": str(fp)}
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(out, ensure_ascii=False))
        return

    headers = [str(c or "").strip() for c in rows[0]]
    hmap = detect_map(headers)

    def get(row, key):
        i = hmap.get(key)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else str(v).strip()

    data_rows = [r for r in rows[1:] if any(c not in (None, "") for c in r)]
    if not data_rows:
        out = {"ok": False, "error": "no_data_rows", "file": str(fp)}
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(out, ensure_ascii=False))
        return

    approved_rows, non_approved_rows = [], []
    status_col_exists = "status" in hmap
    for r in data_rows:
        st = norm(get(r, "status"))
        if status_col_exists:
            if st == "approved":
                approved_rows.append(r)
            else:
                non_approved_rows.append(r)
        else:
            approved_rows.append(r)

    selected_rows = approved_rows
    if not selected_rows:
        out = {
            "ok": False,
            "error": "no_approved_rows",
            "file": str(fp),
            "rows": {
                "total": len(data_rows),
                "approved": 0,
                "ignored_non_approved": len(non_approved_rows),
                "status_column_detected": status_col_exists
            }
        }
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(out, ensure_ascii=False))
        return

    first = selected_rows[0]

    vendor_pd2, vendor_source = find_payment_detail_02_vendor(rows)

    # Prefer direct header-column extraction for Payment Details 01/02, fallback to section scan
    pd01_fallback = collect_section_text(rows, "payment detail 01") or collect_section_text(rows, "payment details 01")
    pd02_fallback = collect_section_text(rows, "payment detail 02") or collect_section_text(rows, "payment details 02")

    def build_recap(row):
        vendor_fallback = get(row, "vendor")
        vendor_final = vendor_pd2 or vendor_fallback
        bill_start = to_date_text(get(row, "billing_start_date"))
        bill_end = to_date_text(get(row, "billing_end_date"))
        due_date = to_date_text(get(row, "due_date"))
        cat = get(row, "category")
        amt = get(row, "amount")
        lines = []
        if cat or amt:
            lines.append({"category": cat, "amount": amt})

        pd01_text = get(row, "payment_details_01") or pd01_fallback
        pd02_text = get(row, "payment_details_02") or pd02_fallback

        header_wht = parse_wht_rate_value(get(row, "wht_rate"))
        parsed_wht, wht_source = extract_wht_rate(pd01_text, ai_cmd=args.wht_ai_cmd)
        wht_rate = header_wht if header_wht is not None else parsed_wht

        return {
            "vendor": vendor_final,
            "vendor_source": vendor_source if vendor_pd2 else ("header_fallback.vendor" if vendor_fallback else vendor_source),
            "bill_number": get(row, "bill_no"),
            "request_no": get(row, "request_no"),
            "bill_date": bill_end,
            "due_date": due_date,
            "location": get(row, "location"),
            "belongs_to": get(row, "belongs_to"),
            "project_type": get(row, "project_type"),
            "reason": get(row, "reason"),
            "payment_detail_01_text": pd01_text,
            "payment_detail_02_text": pd02_text,
            "billing_start_date": bill_start,
            "billing_end_date": bill_end,
            "period_covered": to_period(bill_end),
            "wht_rate": wht_rate if wht_rate is not None else 0.0,
            "wht_amount": get(row, "wht_amount"),
            "wht_parse_source": "header_column" if header_wht is not None else wht_source,
            "vat_flag": get(row, "vat_flag"),
            "lines": lines
        }

    recap = build_recap(first)
    records = []
    for idx, r in enumerate(selected_rows):
        rr = build_recap(r)
        records.append({
            "record_index": idx,
            "recap": rr,
            "missing_required": [k for k in ["vendor", "bill_number", "request_no"] if not rr.get(k)]
        })

    missing = [k for k in ["vendor", "bill_number", "request_no"] if not recap.get(k)]

    kind_rule_hit = {
        "billing_soa_no": bool(recap.get("bill_number")),
        "billing_end_date": bool(recap.get("billing_end_date")),
        "billing_start_date": bool(recap.get("billing_start_date")),
    }
    kind_rule_hit["all_required_for_bill"] = all(kind_rule_hit.values())
    kind = "bill" if kind_rule_hit["all_required_for_bill"] else "unknown"

    out = {
        "ok": True,
        "step": "parse_identify",
        "file": str(fp),
        "sheet": ws.title,
        "header_map": hmap,
        "kind": kind,
        "kind_confidence": "high" if kind == "bill" else "low",
        "kind_rule_hit": kind_rule_hit,
        "needs_user_kind_confirmation": kind != "bill",
        "rows": {
            "total": len(data_rows),
            "approved": len(approved_rows),
            "ignored_non_approved": len(non_approved_rows),
            "status_column_detected": status_col_exists,
            "filter_mode": "strict_approved_only" if status_col_exists else "no_status_column_all_rows_used"
        },
        "recap": recap,
        "records": records,
        "missing_required": missing,
        "parse_hints": {
            "vendor_rule": "Payment detail(s) 02 -> Which supplier",
            "vendor_source_result": recap.get("vendor_source"),
            "pd01_source": "header.payment_details_01_or_section_fallback",
            "pd02_source": "header.payment_details_02_or_section_fallback",
            "bill_date_rule": "bill_date <- billing_end_date",
            "due_date_rule": "due_date <- due_date"
        }
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "out": str(out_path), "file": str(fp), "kind": kind}, ensure_ascii=False))


if __name__ == "__main__":
    main()
